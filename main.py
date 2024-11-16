import os
import openpyxl


def create_events_from_excel():
    wb = openpyxl.load_workbook("events.xlsx")
    sheet = wb.active

    create_calendar()

    events = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9, values_only=True):
        event = create_event(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
        events.append(event)

    with open("calendar.ics", "a") as file:
        for event in events:
            file.write(event)
    print("Events created successfully.")


def create_calendar():
    if 'calendar.ics' in os.listdir():
        os.remove('calendar.ics')

    with open("calendar.ics", "w") as file:
        file.write("BEGIN:VCALENDAR\n")
        file.write("VERSION:2.0\n")
        file.write("PRODID:-//appsheet.com//appsheet 1.0//EN\n")
        file.write("CALSCALE:GREGORIAN\n")
        file.write("METHOD:PUBLISH\n")


def create_event(Title, StartDate, StartTime, EndDate, EndTime, Description, Location, RepeatFrequency, RepeatUntil):
    event = f"BEGIN:VEVENT\n" \
            f"DTSTART:{StartDate}T{StartTime}\n" \
            f"DTEND:{EndDate}T{EndTime}\n" \
            f"SUMMARY:{Title}\n" \
            f"DESCRIPTION:{Description}\n" \
            f"LOCATION:{Location}\n"
    if RepeatFrequency:
        event += f"RRULE:FREQ={RepeatFrequency};UNTIL={RepeatUntil}\n"
    event += "END:VEVENT\n"
    
    return event


def create_new_event():
    Title = input("Enter the title of the event: ")
    StartDate = input("Enter the start date and time (YYYYMMDD): ")
    StartTime = input("Enter the start time (HHMMSS): ")
    EndDate = input("Enter the end date and time (YYYYMMDD): ")
    EndTime = input("Enter the end time (HHMMSS): ")
    Description = input("Enter the event description: ")
    Location = input("Enter the event location: ")

    user_input = input("Do you want to repeat the event? (Y/N): ")
    if user_input == 'Y':
        RepeatFrequency = input("Enter the repeat frequency: (DAILY/WEEKLY/MONTHLY/YEARLY): ")
        RepeatUntil = input("Enter the repeat until date (YYYYMMDDTHHMMSS): ")


    excel_or_ics = input("Do you want to save the event to an excel file or ics file? (1) Excel (2) ICS: ")
    if excel_or_ics == '1':
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append([Title, StartDate, StartTime, EndDate, EndTime, Description, Location, RepeatFrequency, RepeatUntil])
        wb.save("events.xlsx")
    elif excel_or_ics == '2':
        event = create_event(Title, StartDate, StartTime, EndDate, EndTime, Description, Location, RepeatFrequency, RepeatUntil)
        create_calendar()
        with open("calendar.ics", "a") as file:
            file.write(event)


    
if __name__ == "__main__":
    while True:
        user_input = input("What would you like to do? (1/2/3):\n 1. Create events from excel file\n 2. Create a event\n 3. Exit\n")
        if user_input == '1':
            create_events_from_excel()
            break
        elif user_input == '2':
            create_new_event()
            break
        elif user_input == '3':
            break
        else:
            print("Invalid input. Please try again.")



